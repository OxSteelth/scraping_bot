from pathlib import Path
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from PyQt5.QtCore import *
from datetime import datetime, date
from time import sleep
import time
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydrive2.auth import GoogleAuth
from pydrive2.drive import GoogleDrive
from oauth2client.service_account import ServiceAccountCredentials
import pkg_resources
import platform


start_time = time.time()


class Bot:
    running = False
    log_signal = None

    def __init__(self, login_url, email, password, outlet, target_url, date_from, date_to, folder_id):
        # Login
        self.login_url = login_url
        self.email = email
        self.password = password
        self.outlet = outlet

        # Url and date range of the website you want to automatically access
        self.target_url = target_url
        self.date_from = date_from
        self.date_to = date_to

        # other
        self.folder_id = folder_id

    def is_settings_valid(self):
        if not self.login_url or not self.email or not self.password or not self.outlet or not self.target_url or not self.date_from or not self.date_to or not self.folder_id:
            return False
        return True

    def log(self, msg):
        now_time = time.time()
        elapsed_min = int((now_time - start_time) / 60)
        dt_str = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
        output = "{}: [{} min(s)]  {}".format(dt_str, elapsed_min, msg)

        print(output)
        # write to file
        log_file = open("log.txt", "a", encoding='utf-8')
        log_file.write("{}\n".format(output))
        log_file.close()
        # generate signal
        self.log_signal.emit(output)

    def login(self):
        """Login to platform"""
        self.driver.find_element('name', 'account').send_keys(self.email)
        self.driver.find_element('name', 'password').send_keys(self.password)
        self.driver.find_element('xpath', '//*[@id="cForm_row_system_signIn_outlet"]/div/div[1]').click()
        sleep(1)
        self.driver.find_element('xpath', '//*[@id="cForm_row_system_signIn_outlet"]/div/div[2]/div[1]/input').send_keys(self.outlet)
        sleep(1)
        self.driver.find_element('xpath', '//*[@id="cForm_row_system_signIn_outlet"]/div/div[2]/div[2]/ul/li').click()
        self.driver.find_element('xpath', '//*[@id="cForm_action_system_signIn"]/button[1]').click()

    def select_collection(self):
        """Select collection to extract data"""
        while True:
            try:
                self.driver.get(self.target_url)
                sleep(3)
                break
            except:
                continue

    def open_report(self):
        """Open report to extract data"""
        self.driver.find_element('id', 'cForm_reporter_reportForm_dateFrom').send_keys(self.date_from)
        self.driver.find_element('id', 'cForm_reporter_reportForm_dateTo').send_keys(self.date_to)
        self.driver.find_element('id', 'siteTitle').click()
        self.driver.find_element('xpath', '//*[@id="cForm_row_reporter_reportForm_outlet"]/ul/li[1]').click()
        self.driver.find_element('xpath', '//*[@id="cForm_action_reporter_reportForm"]/button[1]').click()

    def upload_file_to_gdrive(self, folder_id, path, file_name):
        gauth = GoogleAuth()
        # NOTE: if you are getting storage quota exceeded error, create a new service account, and give that service account permission to access the folder and replace the google_credentials.
        gauth.credentials = ServiceAccountCredentials.from_json_keyfile_name(pkg_resources.resource_filename(__name__, "credentials.json"), scopes=['https://www.googleapis.com/auth/drive'])

        drive = GoogleDrive(gauth)

        file = drive.CreateFile({'parents': [{"id": folder_id}], 'title': file_name})

        file.SetContentFile(path + '/' + file_name)
        file.Upload()

    def extract(self, m_workbook: openpyxl.Workbook):
        """Extract data one by one"""
        base_dir = Path(__file__).resolve().parent
        # Driver of the browser you use
        try:
            os_type = platform.system()
            if os_type == 'Windows':
                self.driver = webdriver.Chrome("chromedriver.exe")
            elif os_type == 'Linux':
                self.driver = webdriver.Chrome("{0}/chromedriver".format(base_dir))
            elif os_type == 'Darwin':
                chrome_options = webdriver.ChromeOptions()
                chrome_options.add_argument("--remote-debugging-port=9222")
                self.driver = webdriver.Chrome("{0}/chromedriver".format(base_dir), chrome_options=chrome_options)
            else:
                self.log("Not recognized OS type.")
                return
        except Exception as e:
            self.log(str(e))
            return

        try:
            self.browse = self.driver.get(self.login_url)
        except Exception as e:
            self.log(str(e))
            self.driver.close()
            return

        # start log
        self.log("Bot started ...")
        # Login to platform
        self.login()
        sleep(5)

        # Select collection
        self.select_collection()

        # Open report
        self.open_report()

        # wait until tables are loaded
        WebDriverWait(self.driver, 60).until(expected_conditions.presence_of_element_located((By.XPATH, '//*[@id="siteContent"]/form[@class="cTable"]')))
        sleep(1)

        self.log("Retrieving data ...")

        # find target forms
        forms = self.driver.find_elements('xpath', '//*[@id="siteContent"]/form[@method="post"]')
        # remove summary form
        forms.pop(0)

        fields = []
        # retrieve the fields
        columns = forms[0].find_elements('xpath', './/table/thead/tr/th')
        for column in columns:
            try:
                head = column.find_element('xpath', './/a').text.strip()
                fields.append(head)
                if head == 'Payment types':
                    fields.append('OUTLET')
            except:
                pass
        fields.append('Customer link')
        fields.append('Invoice link')

        data = []
        for form in forms:
            outlet = form.find_element('xpath', './/h1').text.strip()

            # extract data
            rows = form.find_elements('xpath', './/table/tbody/tr')
            for row in rows:
                datum = {'OUTLET': outlet}
                i = 0

                cells = row.find_elements('xpath', './/td')
                for cell in cells:
                    # skip this field that doesn't exist in table columns
                    if fields[i] == 'OUTLET':
                        i += 1

                    try:
                        datum[fields[i]] = cell.find_element('xpath', './/div').text
                    except:
                        try:
                            datum[fields[i]] = cell.find_element('xpath', './/a').get_attribute('href')
                        except:
                            datum[fields[i]] = ''
                    i += 1

                data.append(datum)

            # remove last row where total amounts are put
            data.pop()

        # close driver
        self.driver.close()

        self.log("Saving extracts ...")

        sheet = m_workbook.active

        # Create a bold font
        bold_font = Font(bold=True)

        # Create a small font
        small_font = Font(name='Trebuchet MS', size='8')

        # Create a large font
        large_font = Font(name='Aptos Narrow', size='12')

        # Create a link font
        link_font = Font(name='Aptos Narrow', size='12', color='467886')

        # Create a pink fill
        pinkFill = PatternFill(start_color='F2CEEF', end_color='F2CEEF', fill_type='solid')

        # Create a green fill
        greenFill = PatternFill(start_color='DAF2D0', end_color='DAF2D0', fill_type='solid')

        # Create a blue fill
        blueFill = PatternFill(start_color='CAEDFB', end_color='CAEDFB', fill_type='solid')

        # set the height of the row 
        sheet.row_dimensions[1].height = 64

        fields.insert(0, '')

        i = 0
        for field in fields:
            if field in ['Insurance Num.', 'Remarks']:
                continue
            i += 1

            sheet.column_dimensions[get_column_letter(i)].width = 16

            cell = sheet.cell(row=1, column=i)
            if field not in ['Customer link', 'Invoice link']:
                cell.value = field
            cell.alignment = Alignment(wrap_text=True)
            cell.font = bold_font
            if field == 'OUTLET':
                cell.fill = greenFill
            else:
                cell.fill = pinkFill

        fields.pop(0)

        i = 1
        for datum in data:
            i += 1
            j = 1

            cell = sheet.cell(row=i, column=j)
            cell.value = datum['Outlet']
            cell.font = large_font

            for field in fields:
                if field in ['Insurance Num.', 'Remarks']:
                    continue
                j += 1
                cell = sheet.cell(row=i, column=j)
                if field in ['Customer link', 'Invoice link']:
                    cell.value = field.split(' ')[0]
                    cell.hyperlink = datum[field]
                    cell.style = 'Hyperlink'
                    cell.font = link_font
                else:
                    cell.value = datum[field]
                    cell.font = small_font
                    if field == 'OUTLET':
                        cell.fill = blueFill

        data_dir = Path.joinpath(base_dir, 'data')
        path = str(data_dir) + '/' + datetime.today().strftime('%Y/%m/%d')
        if not os.path.exists(path):
            os.makedirs(path)
        file_name = '{0}.xlsx'.format(date.today().strftime("%Y-%m-%d"))
        m_workbook.save(path + '/' + file_name)

        try:
            self.upload_file_to_gdrive(self.folder_id, path, file_name)
            self.log("Uploaded {0}".format(file_name))
        except Exception as e:
            self.log(str(e))

        self.log("Finished.")
